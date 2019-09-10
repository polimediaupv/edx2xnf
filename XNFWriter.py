# coding=utf-8
'''
from xlutils.copy import copy
from xlrd import *
from xml.etree import ElementTree as ET
from ast import literal_eval as make_tuple
'''
import os
import codecs
import shutil
import datetime
from lxml import etree,html
import argparse
from dateutil import parser as dateparser
import pdb
import tarfile
import openpyxl
import json
import sys
reload(sys)
sys.setdefaultencoding('utf8')

def untar(cfg):    
    shutil.rmtree(cfg['UNCOMPRESSFOLDER'])
    tar = tarfile.open(cfg['ORIGINFOLDER']+cfg['FILENAME'])    
    tar.extractall(cfg['UNCOMPRESSFOLDER'])
    coursefoldername = tar.getmembers()[0].name
    tar.close()
    return coursefoldername

def getCourseMetadata(course_url_name):    
    root = etree.parse(course_url_name)
    attributes = root.getroot().attrib
    return attributes

def fixencoding(text):
    result = text.encode('latin-1').decode('utf-8')
    return result
    



def updateDatosgenerales(originFolder,course_attributes,cfg,wb):
    if 'course' in course_attributes:
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["NAME"]] = course_attributes['course']        
    if 'start' in course_attributes:
        #maybe need to format start date             
        startdate = dateparser.parse(course_attributes['start'].replace('"',''))
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["STARTDATE"]].value = datetime.datetime.strptime(startdate.strftime("%d/%m/%Y"), "%d/%m/%Y")
    if 'display_name' in course_attributes:
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["DISPLAYNAME"]].value = course_attributes['display_name']
    if 'url_name' in course_attributes:
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["NUMBER"]].value = course_attributes['url_name']
    ## need to parse course overview to get rest of the data ( ͡° ͜ʖ ͡°) 
    # TODO : Check if this sections exists prior to parse them to avoid problems (:
    htmltree = html.parse(originFolder+ '/about/overview.html')
    if len(htmltree.getroot().cssselect('section.info'))>0:
        info = htmltree.getroot().cssselect('section.info')[0]    
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["COURSEDESCRIPTION"]] = fixencoding(info.text_content())
    if len(htmltree.getroot().cssselect('section.about'))>0:    
        objetives = htmltree.getroot().cssselect('section.about')[0]
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["COURSEOBJETIVES"]] = fixencoding(objetives.text_content())
    if len(htmltree.getroot().cssselect('section.prerequisites'))>0:
        prerequisites = htmltree.getroot().cssselect('section.prerequisites')[0]
        wb['DatosGenerales'][cfg["DATOSGENERALES"]["COURSEPREREQUISITES"]] = fixencoding(prerequisites.text_content())
    if len(htmltree.getroot().cssselect('section.course-staff')[0].cssselect('article'))>0:
        courseStaff = htmltree.getroot().cssselect('section.course-staff')[0].cssselect('article')
        teacherNames=''
        TeacherBioRow=cfg["DATOSGENERALES"]['BIOSTARTINGROW']
        
        for teacher in courseStaff:
            if len(teacher.cssselect('h3'))>0:
                teacherNames += fixencoding(teacher.cssselect('h3')[0].text_content())+';'
            else:
                teacherNames += 'se esperaba etiqueta h3;'
            teacherBio = fixencoding(teacher.cssselect('p')[-1].text_content())
            wb['DatosGenerales']["A"+unicode(TeacherBioRow)] = unicode(teacherBio)
            TeacherBioRow+=1        
        teacherNames = teacherNames[:-1]    
        wb['DatosGenerales'][cfg["DATOSGENERALES"]['TEACHERNAMES']]=teacherNames
        wb['DatosGenerales']._tables[0].ref = 'A5:A' + unicode(TeacherBioRow-1)    
    overview = html.parse(originFolder+ '/about/overview.html')
    return wb

#need to move this part so we can make it recursive calls because
#for some reason there are problems inside problems what makes 
#no sense in the original design
def writeProblemType(problemXML,wb,problemRow):
    for elem in problemXML.getchildren():        
        if elem.tag=='p':
            wb['Problemas'][cfg["PROBLEMAS"]['ENUNCIADOCOL'] + unicode(problemRow)].value = etree.tostring(elem,pretty_print=True)        
        if elem.tag=='multiplechoiceresponse':
            wb['Problemas'][cfg["PROBLEMAS"]['TIPOCOL'] + unicode(problemRow)].value = 'MultiChoice'
            answerCols=cfg["PROBLEMAS"]['ANSWERCOLS']
            answer = 0         
            for group in elem.getchildren():
                if group.tag=='choicegroup':
                    for choice in group.getchildren():
                        if 'correct' in choice.attrib:
                            if choice.attrib['correct']=='true':                                
                                wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value=unicode(answer+1)
                        wb['Problemas'][answerCols[answer] + unicode(problemRow)].value = etree.tostring(choice, encoding='unicode', with_tail=False,method='text')
                        answer += 1
        if elem.tag=='choiceresponse':
            wb['Problemas'][cfg["PROBLEMAS"]['TIPOCOL']  + unicode(problemRow)].value = 'CheckBox'
            answerCols=cfg["PROBLEMAS"]['ANSWERCOLS'] 
            answer = 0  
            for group in elem.getchildren():
                if group.tag=='checkboxgroup':
                    for choice in group.getchildren():
                        if 'correct' in choice.attrib:
                            if choice.attrib['correct']=='true':                                      
                                if wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value != None:
                                    wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value+=unicode(answer+1)+';'
                                else:
                                    wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value=unicode(answer+1)+';'
                        wb['Problemas'][answerCols[answer] + unicode(problemRow)].value = etree.tostring(choice, encoding='unicode', with_tail=False,method='text')
                        answer += 1
            
            if wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value != None:
                wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value=wb['Problemas'][cfg["PROBLEMAS"]['CORRECTCOL'] + unicode(problemRow)].value[:-1]
            
        if elem.tag=='problem':
            writeProblemType(elem,wb,problemRow)            
    return wb


def writeproblem(originFolder,cfg,wb,attributes,lesson):
    problemXML = etree.parse(originFolder+'/problem/'+attributes['url_name']+'.xml').getroot()
    if wb['Problemas'].max_row==2 and wb['Problemas']['B2'].value==None:
        problemRow=2
    else:
        problemRow = wb['Problemas'].max_row +1 
    
    wb['Problemas'][cfg["PROBLEMAS"]['NOMBRESECCIONCOL'] + unicode(problemRow)].value=lesson['seccion']
    wb['Problemas'][cfg["PROBLEMAS"]['NOMBRESUBSECCIONCOL'] + unicode(problemRow)].value=lesson['subseccion']
    wb['Problemas'][cfg["PROBLEMAS"]['TITULOCOL'] + unicode(problemRow)].value=lesson['lesson']
    if 'max_attempts' in problemXML.attrib:
        wb['Problemas'][cfg["PROBLEMAS"]['MAXATTEMPTSCOL'] + unicode(problemRow)].value=problemXML.attrib['max_attempts']
    if 'weight' in problemXML.attrib:
        wb['Problemas'][cfg["PROBLEMAS"]['WEIGHTCOL'] + unicode(problemRow)].value=problemXML.attrib['weight']
    if 'showanswer' in problemXML.attrib:
        if problemXML.attrib['showanswer'] == 'Answered':
            wb['Problemas'][cfg["PROBLEMAS"]['SHOWANSWERCOL'] + unicode(problemRow)].value='Respondida'
        if problemXML.attrib['showanswer'] == 'Allways':
            wb['Problemas'][cfg["PROBLEMAS"]['SHOWANSWERCOL'] + unicode(problemRow)].value='Siempre'
        if problemXML.attrib['showanswer'] == 'Never':
            wb['Problemas'][cfg["PROBLEMAS"]['SHOWANSWERCOL'] + unicode(problemRow)].value='Nunca'
        if problemXML.attrib['showanswer'] == 'Ended':
            wb['Problemas'][cfg["PROBLEMAS"]['SHOWANSWERCOL'] + unicode(problemRow)].value='Terminada'
    #J randomice
    wb = writeProblemType(problemXML,wb,problemRow)    
    wb['Problemas'][cfg["PROBLEMAS"]['FORMULASECCIONCOL'] + unicode(problemRow)].value=u'=INDEX(Unidades[#Data],MATCH(Problemas[[#This Row],[Secci\xf3n]],Unidades[selectorSeccion],0),2)'
    wb['Problemas'][cfg["PROBLEMAS"]['FORMULASUBSECCIONCOL']+ unicode(problemRow)].value=u'=INDEX(Unidades[],MATCH(Problemas[[#This Row],[SubSecci\xf3n]],Unidades[selectorSubSeccion],0),4)'
    wb['Problemas'][cfg["PROBLEMAS"]['FORMULALESSONCOL'] + unicode(problemRow)].value=u'=INDEX(Leccion[],MATCH(Problemas[[#This Row],[Lecci\xf3n]],Leccion[selectorLeccion],0),5)'

    wb['Problemas']._tables[0].ref = cfg["PROBLEMAS"]['TABLEDEF'] + unicode(problemRow)      
    for val in wb['Problemas'].data_validations.dataValidation:        
        val.sqref=openpyxl.worksheet.cell_range.MultiCellRange([unicode(val.sqref)[:1]+'2:'+unicode(val.sqref)[:1]+unicode(problemRow)])       
    return wb
    

def writeSequential(originFolder,attributes,cfg,wb,chapterAttrib,numChapter,numSequential):
    #calculamos row con maxrow y 2
    if wb['Unidades'].max_row==2 and wb['Unidades']['C2'].value==None:
        row=2
    else:
        row = wb['Unidades'].max_row +1

    sequentialXML =etree.parse(originFolder+'/sequential/'+attributes['url_name']+'.xml').getroot()
    wb['Unidades'][cfg["UNIDADES"]['NOMBRESECCIONCOL'] + unicode(row)]=chapterAttrib['display_name']
    wb['Unidades'][cfg["UNIDADES"]['NOMBRESUBSECCIONCOL']+ unicode(row)]=sequentialXML.attrib['display_name']
    if 'start' in chapterAttrib:        
        startdate =  dateparser.parse(chapterAttrib['start'].replace('"',''))                      
        wb['Unidades'][cfg["UNIDADES"]['STARTDATECOL']+ unicode(row)].value =datetime.datetime.strptime(startdate.strftime("%d/%m/%Y"), "%d/%m/%Y")
    else:
        wb['Unidades'][cfg["UNIDADES"]['STARTDATECOL']+ unicode(row)]=''
    if 'end' in chapterAttrib:
        enddate =  dateparser.parse(chapterAttrib['end'].replace('"',''))
        wb['Unidades'][cfg["UNIDADES"]['ENDDATECOL']+ unicode(row)].value = datetime.datetime.strptime(enddate.strftime("%d/%m/%Y"), "%d/%m/%Y")
    else:
        wb['Unidades'][cfg["UNIDADES"]['ENDDATECOL']+ unicode(row)]=''
    #unset tasktype this functionality does not work in edx platform so can't be backported
    wb['Unidades'][cfg["UNIDADES"]['TASKTYPECOL']+ unicode(row)]=''

    wb['Unidades']._tables[0].ref = 'A1:L' + unicode(row)   
    #we copy the formulas
    if row > 2:
        #pdb.set_trace()
        wb['Unidades']['B' + unicode(row)].value = '=IFERROR(IF(C'+ unicode(row-1)+ '=C'+ unicode(row)+ ',B'+ unicode(row-1)+ ',B'+ unicode(row-1)+ '+1),1)'
        wb['Unidades']['D' + unicode(row)].value = '=IF(B'+ unicode(row)+ '=B'+ unicode(row-1)+ ',D'+ unicode(row-1)+ '+1,1)'
        wb['Unidades']['K' + unicode(row)].value = wb['Unidades']['K2'].value
        wb['Unidades']['L' + unicode(row)].value = wb['Unidades']['L2'].value
        #pdb.set_trace()
    #this calls updateLesson
    lesson = {"seccion":unicode(numChapter)+":"+chapterAttrib['display_name'],
            "subseccion":unicode(numChapter)+":"+unicode(numSequential)+":"+sequentialXML.attrib['display_name'],            
            "titulo":"",
            "objetivos":"",
            "video":"",
            "resumen":"",
            "forum":""
    }
    lessonNumber = 1
    for vertical in sequentialXML.cssselect('vertical'):                
        verticalXML =etree.parse(originFolder+'/vertical/'+vertical.attrib['url_name']+'.xml').getroot()                
        if 'display_name' in verticalXML.attrib:            
            if lesson['titulo']!='' and lesson['titulo']!= verticalXML.attrib['display_name'] and verticalXML.attrib['display_name']!='':                
                if wb['Leccion'].max_row==2 and wb['Leccion']['B2'].value==None:
                    lessonRow=2
                else:
                    lessonRow = wb['Leccion'].max_row +1 
                
                wb['Leccion'][cfg['LECCION']['NOMBRESECCIONCOL'] + unicode(lessonRow)].value=lesson['seccion']
                wb['Leccion'][cfg['LECCION']['NOMBRESUBSECCIONCOL'] + unicode(lessonRow)].value=lesson['subseccion']
                wb['Leccion'][cfg['LECCION']['TITULOCOL'] + unicode(lessonRow)].value=lesson['titulo']
                wb['Leccion'][cfg['LECCION']['OBJETIVOSCOL'] + unicode(lessonRow)].value=lesson['objetivos']
                wb['Leccion'][cfg['LECCION']['VIDEOCOL'] + unicode(lessonRow)].value=lesson['video']
                wb['Leccion'][cfg['LECCION']['RESUMENCOL'] + unicode(lessonRow)].value=lesson['resumen']
                #copy the formulas    
                wb['Leccion'][cfg['LECCION']['FORMULASECCIONCOL'] + unicode(lessonRow)].value=u'=INDEX(Unidades[],MATCH(Leccion[[#This Row],[Secci\xf3n]],Unidades[selectorSeccion],0),2)'
                wb['Leccion'][cfg['LECCION']['FORMULASUBSECCIONCOL'] + unicode(lessonRow)].value=u'=INDEX(Unidades[],MATCH(Leccion[[#This Row],[SubSecci\xf3n]],Unidades[selectorSubSeccion],0),4)'
                wb['Leccion'][cfg['LECCION']['FORMULALESSONCOL'] + unicode(lessonRow)].value='=IFERROR((IF(AND(A' + unicode(lessonRow-1) + '=A' + unicode(lessonRow) + ',C' + unicode(lessonRow-1) + '=C' + unicode(lessonRow) + '),E' + unicode(lessonRow-1) + '+1,1)),"")'
                wb['Leccion'][cfg['LECCION']['FORMULACONCATCOL'] + unicode(lessonRow)].value=u'=CONCATENATE(INDIRECT("Leccion[idSeccion]"),":",INDIRECT("Leccion[idSubSeccion]"),":",INDIRECT("Leccion[Lecci\xf3n]"),":",INDIRECT("Leccion[Titulo]"))'

                wb['Leccion']._tables[0].ref = cfg['LECCION']['TABLEDEF'] + unicode(lessonRow)   
                lessonNumber+=1

            lesson['titulo']= verticalXML.attrib['display_name']
            lesson["lesson"]=unicode(numChapter)+":"+unicode(numSequential)+":" + unicode(lessonNumber) + ":"+lesson['titulo']
            for child in verticalXML.getchildren():                
                if child.tag=="html":
                    htmlFiletext = open(originFolder+'/html/'+child.attrib['url_name']+'.html', 'r').read() 
                    if lesson['video']!="":
                        lesson['resumen']=htmlFiletext
                    else:
                        lesson['objetivos']=htmlFiletext                    
                if child.tag=="video":
                    videoXML = etree.parse(originFolder+'/video/'+child.attrib['url_name']+'.xml').getroot()
                    lesson["video"]=videoXML.attrib['youtube_id_1_0']
                if child.tag=="discussion":
                    lesson["forum"]="1"
                if child.tag=="problem":
                    writeproblem(originFolder,cfg,wb,child.attrib,lesson)
                    
    if wb['Leccion'].max_row==2 and wb['Leccion']['B2'].value==None:
        lessonRow=2
    else:
        lessonRow = wb['Leccion'].max_row +1 

    #TODO - change letters with columnnames in cfg file
    wb['Leccion'][cfg['LECCION']['NOMBRESECCIONCOL'] + unicode(lessonRow)].value=lesson['seccion']
    wb['Leccion'][cfg['LECCION']['NOMBRESUBSECCIONCOL'] + unicode(lessonRow)].value=lesson['subseccion']
    wb['Leccion'][cfg['LECCION']['TITULOCOL'] + unicode(lessonRow)].value=lesson['titulo']
    wb['Leccion'][cfg['LECCION']['OBJETIVOSCOL'] + unicode(lessonRow)].value=lesson['objetivos']
    wb['Leccion'][cfg['LECCION']['VIDEOCOL'] + unicode(lessonRow)].value=lesson['video']
    wb['Leccion'][cfg['LECCION']['RESUMENCOL'] + unicode(lessonRow)].value=lesson['resumen']
    #copy the formulas    
    wb['Leccion'][cfg['LECCION']['FORMULASECCIONCOL'] + unicode(lessonRow)].value=u'=INDEX(Unidades[],MATCH(Leccion[[#This Row],[Secci\xf3n]],Unidades[selectorSeccion],0),2)'
    wb['Leccion'][cfg['LECCION']['FORMULASUBSECCIONCOL'] + unicode(lessonRow)].value=u'=INDEX(Unidades[],MATCH(Leccion[[#This Row],[SubSecci\xf3n]],Unidades[selectorSubSeccion],0),4)'
    wb['Leccion'][cfg['LECCION']['FORMULALESSONCOL'] + unicode(lessonRow)].value='=IFERROR((IF(AND(A' + unicode(lessonRow-1) + '=A' + unicode(lessonRow) + ',C' + unicode(lessonRow-1) + '=C' + unicode(lessonRow) + '),E' + unicode(lessonRow-1) + '+1,1)),"")'
    wb['Leccion'][cfg['LECCION']['FORMULACONCATCOL'] + unicode(lessonRow)].value=u'=CONCATENATE(INDIRECT("Leccion[idSeccion]"),":",INDIRECT("Leccion[idSubSeccion]"),":",INDIRECT("Leccion[Lecci\xf3n]"),":",INDIRECT("Leccion[Titulo]"))'

    wb['Leccion']._tables[0].ref = cfg['LECCION']['TABLEDEF'] + unicode(lessonRow)   
    
    for val in wb['Leccion'].data_validations.dataValidation:
        val.sqref=openpyxl.worksheet.cell_range.MultiCellRange([unicode(val.sqref)[:1]+'2:'+unicode(val.sqref)[:1]+unicode(lessonRow)])        
    return wb

def updateUnidades(originFolder,attributes,cfg,wb):
    '''
    para cada chapter en el course.xml mirar cuantos sequential tiene dentro del xml y añadir una entrada en la columna C de la tabla unidades
    luego mirar el archivo xml que esta enlazado en cada uno de los sequentials y sacar el displayname de la subsection    
    '''
    courseXML =etree.parse(originFolder+'/course/'+attributes['url_name']+'.xml').getroot()    
    numChapter = 1
    for chapter in courseXML.cssselect('chapter'):
        chapterXML = etree.parse(originFolder+'/chapter/'+chapter.attrib['url_name']+'.xml').getroot()
        numSequential = 1
        for sequential in chapterXML.cssselect('sequential'):
            wb = writeSequential(originFolder,sequential.attrib,cfg,wb,chapterXML.attrib,numChapter,numSequential)            
            numSequential +=1
        numChapter +=1
        
    return wb




    


def generate_XNF(cfg):    
    originFolder = cfg['UNCOMPRESSFOLDER'] + '/' + untar(cfg)
    if 'xlsm' in cfg['TARGETNAME']:
        wb = openpyxl.load_workbook(cfg['TEMPLATEPATH'],keep_vba=True)
    else:
        wb = openpyxl.load_workbook(cfg['TEMPLATEPATH'])
    #get course num and course title from xml
    for f in os.listdir(originFolder):
        if '.xml' in f:
            root = etree.parse(originFolder+ '/' + f )
            attributes = root.getroot().attrib             
            print("Starting generation of : " + attributes['course'])
            courseAttributes = getCourseMetadata(originFolder+'/course/'+attributes['url_name']+'.xml')
            for key in courseAttributes:
                if key not in attributes:
                    attributes[key]=courseAttributes[key]
            wb = updateDatosgenerales(originFolder,attributes,cfg,wb)    
            wb = updateUnidades(originFolder,attributes,cfg,wb)    
            
    wb.save(cfg['TARGETNAME'])   



parser = argparse.ArgumentParser(description='XNF generator')
parser.add_argument('--f', dest='file',default=None,help='tar.gz file containing the original course, if not declared will use the one declared in the cfg file')
parser.add_argument('--p', dest='path',default=None,help='the origin folder where the original tag is located, if not declared will use the one declared in the cfg file')
args = parser.parse_args()
cfg = json.loads(open("conf.json").read())
if args.file!=None:    
    cfg["TEMPLATEPATH"]=args.file
if args.path!=None:    
    cfg['ORIGINFOLDER']=args.path
generate_XNF(cfg)

